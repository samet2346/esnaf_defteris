import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from customers.models import Musteri
from jobs.models import IsKaydi


class ExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Musteriler"
        ws1.append(["ID", "Ad Soyad", "Telefon", "Adres", "Toplam Harcama", "Kalan Borc"])
        for m in Musteri.objects.filter(user=user):
            ws1.append([m.id, m.ad_soyad, m.telefon, m.adres or "", float(m.toplam_harcama), float(m.kalan_borc)])

        ws2 = wb.create_sheet("Is Kayitlari")
        ws2.append(["Talep No", "Musteri", "Yapilan Is", "Ucret", "Odenen Tutar", "Durum", "Tarih"])
        for j in IsKaydi.objects.filter(user=user).select_related("musteri"):
            ws2.append([
                j.talep_no,
                j.musteri.ad_soyad,
                j.yapilan_is,
                float(j.ucret),
                float(j.odenen_tutar),
                j.durum,
                j.tarih.strftime("%Y-%m-%d %H:%M"),
            ])

        for ws in (ws1, ws2):
            for i, column_cells in enumerate(ws.columns, 1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=esnaf_defteri_verileri.xlsx"
        wb.save(response)
        return response


class ExcelImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"detail": "Dosya bulunamadi."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception:
            return Response({"detail": "Gecerli bir Excel dosyasi yukleyin."}, status=status.HTTP_400_BAD_REQUEST)

        if "Musteriler" not in wb.sheetnames:
            return Response({"detail": "Dosyada Musteriler sayfasi bulunamadi."}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb["Musteriler"]
        eklenen = 0
        atlanan = 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[1]:
                continue
            ad_soyad = str(row[1]).strip()
            telefon = str(row[2]).strip() if row[2] else ""
            adres = str(row[3]).strip() if len(row) > 3 and row[3] else ""

            if not ad_soyad or not telefon:
                atlanan += 1
                continue

            _, created = Musteri.objects.get_or_create(
                user=user,
                telefon=telefon,
                defaults={"ad_soyad": ad_soyad, "adres": adres},
            )
            if created:
                eklenen += 1
            else:
                atlanan += 1

        return Response({
            "detail": f"{eklenen} musteri eklendi, {atlanan} kayit atlandi (zaten mevcut veya eksik bilgi).",
            "eklenen": eklenen,
            "atlanan": atlanan,
        })
